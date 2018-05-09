#!/usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'orleven'

import argparse
import os

from openpyxl import Workbook
'''
    批量 dic行的log导入到excal
    python diclogtoexcal.py -M   生成多文件
    python diclogtoexcal.py -O 文件名  生成单文件
'''


titleList = []
fileList = []


def txtToExcal_1(filename):
    with open (filename,'rb') as f:
        myfilename =  filename + '.xlsx'
        book = Workbook()
        ws = book.active
        i,j = 1,1
        title = []
        for line in f.readlines():
            mydic = eval(line)
            i = i+1
            for key in mydic:
                if key not in title:
                    title.append(key)
                    ws.cell(row = 1  ,column  = len(title)).value = key
                try:
                    ws.cell(row =i  ,column  = title.index(key)+1).value = mydic[key].encode('utf-8')
                except:
                    ws.cell(row =i  ,column  = title.index(key)+1).value = "None"
        book.save(myfilename)


def txtToExcal_2(filename,i,ws,out):
    with open (filename,'rb') as f:
        for line in f.readlines():
            mydic = eval(line.decode('utf-8'))
            i = i+1
            for key in mydic:
                if key not in titleList:
                    titleList.append(key)
                    ws.cell(row = 1  ,column  = len(titleList)).value = key
                try:
                    ws.cell(row =i  ,column  = titleList.index(key)+1).value = mydic[key].decode('utf-8')
                except:
                    ws.cell(row =i  ,column  = titleList.index(key)+1).value = "None"
    return i

def deal(myfile,out,mode):
    if not mode :
        myfilename =  out + '.xlsx'
        book = Workbook()
        ws = book.active
        i = 1
        for filename in fileList:
            print "[+] Dealing file (" + filename +")..."
            i = txtToExcal_2(filename,i,ws,out)
        book.save(myfilename)
    else:
        for filename in fileList:
            print "[+] Dealing file (" + filename +")..."
            txtToExcal_1(filename)



def argSet(parser):
    parser.add_argument("-F", "--filename",type=str, help="Load file e.g. a.html,/tmp/test/", default=None,required=False)
   
    parser.add_argument("-O", "--out",type=str,help="Output file", default="result")
    parser.add_argument("-M", "--mode",action='store_true',help="one file or mul file", default=False)
    return parser

def handle(args):
    filename = args.filename
    filename = './'
    mode = args.mode
    out = args.out

    print "[+] Start run."
    if filename != None:
        if os.path.isfile(filename):
            fileList.append(myfile)
        elif os.path.isdir(filename):
            print "[+] Dealing dir (" + filename +")" + "..."
            for myfile in os.listdir(filename) :
                if os.path.splitext(myfile)[1] == '.log':  
                    fileList.append(myfile)
                    
        else:
            print "[+] The path is not exist!"
    else:
        print "[-] The path is not null!"
    deal(myfile,out,mode)
    print "[+] End of run."
        
if __name__=='__main__':
    parser = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter)
    parser = argSet(parser)
    args = parser.parse_args()
    handle(args)
    
