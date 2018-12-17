#!/usr/bin/env python
# -*- coding: utf-8 -*-
__author__ = 'orleven'

from bs4 import BeautifulSoup
from openpyxl import Workbook
import argparse

def argSet(parser):
    parser.add_argument("-F", "--filename",type=str, help="", default=None,required=False)
    return parser

def deal(filename):
    print "[+] Dealing file ...!"
    headers = [u"序号", u"漏洞名称",  u"影响资产", u"风险等级", u'CVSS', u'CVE', u"漏洞描述", u"解决方案", u"html报告中序号"]
    book = Workbook()
    ws = book.active
    htmlfile = open(filename, 'rb')
    htmlhandle = htmlfile.read()
    soup = BeautifulSoup(htmlhandle, 'html.parser')
    div =  soup.find(id="searchBox")
    for i in range(0, len(headers)):
        ws.cell(row=1, column=i + 1, value=headers[i])
    i = 2
    for item in div.find_all('li'):
        div = item.find('div',class_='fold-title')
        if div == None:
            continue
        no =  div.span.get_text()
        name = div.div.get_text()

        tds =  item.find_all('td')
        level = tds[1].span.get_text()
        cvss = tds[2].get_text()
        cve = tds[3].get_text()
        describe = tds[4].p.get_text()
        idea = tds[5].get_text().strip()
        
        for assets in  tds[0].find_all('a'):
            asset = assets.get_text()
            ws.cell(row=i, column= 1, value=i-1)
            ws.cell(row=i, column= 2, value=name)
            ws.cell(row=i, column= 3, value=asset)
            ws.cell(row=i, column= 4, value=level)
            ws.cell(row=i, column= 5, value=cvss)
            ws.cell(row=i, column= 6, value=cve)
            ws.cell(row=i, column= 7, value=describe)
            ws.cell(row=i, column= 8, value=idea)
            ws.cell(row=i, column= 9, value=no)
            i+=1
    print "[+] Saving file..."
    book.save(filename+'.xlsx')
    print "[+] Successful!"
    
def handle(args):
    filename = args.filename
    if filename != None :
        deal(filename)
    else:
        print '[-] Please input file or lookup help!'
    
if __name__=='__main__':
    parser = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter)
    parser = argSet(parser)
    args = parser.parse_args()
    handle(args)
    
    
